import openpyxl
import configparser
from openpyxl import Workbook
from openpyxl.drawing.image import Image as xlImage
from minilog import Minilog

log = Minilog()
cfg = configparser.ConfigParser()
cfg.read('config.ini')


def add_to_excel(uid: str, fio: str, doljnost: str, roomname: str, status: str, data: str, time: str,
                 foto_path: str, discipline_status: dict, excel_file: str):
    """
    Записывает переданные значения в файл Excel.

    Параметры:
        uid (str): UID сотрудника.
        fio (str): ФИО сотрудника.
        doljnost (str): Должность сотрудника.
        roomname (str): Название помещения.
        status (str): Статус (приход/уход).
        data (str): Дата.
        time (str): Время.
        foto_path (str): Путь к фотографии.
        opozdanie (str): Информация об опоздании.
        ran_uhod (str): Информация о раннем уходе.
        progul (str): Информация о прогуле.
        zaderjka (str): Информация о задержке на работе.
        excel_file (str): Путь к файлу Excel.
    """
    try:

        # Попытка загрузить существующий файл Excel или создать новый
        wb = openpyxl.load_workbook(excel_file)
        log.write_log("Загрузить таблицу удалось!")
    except FileNotFoundError:
        log.write_error(f"Файл по указанному пути -> {excel_file} - не найден!")
        wb = Workbook()
    sheet = wb.active

    # Добавление заголовков, если они ещё не добавлены
    if sheet.max_row == 1:
        headers = ["UID", "ФИО", "Должность", "Помещение", "Статус", "Дата", "Время", "Опоздание", "Ранний уход",
                   "Прогул", "Задержка", "Фотография"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

    row = sheet.max_row + 1
    # Добавление данных
    sheet.cell(row=row, column=1, value=uid)
    sheet.cell(row=row, column=2, value=fio)
    sheet.cell(row=row, column=3, value=doljnost)
    sheet.cell(row=row, column=4, value=roomname)
    sheet.cell(row=row, column=5, value=status)
    sheet.cell(row=row, column=6, value=data)
    sheet.cell(row=row, column=7, value=time)
    sheet.cell(row=row, column=8, value=discipline_status['opoz'])
    sheet.cell(row=row, column=9, value=discipline_status['rann'])
    sheet.cell(row=row, column=10, value=discipline_status['prog'])
    sheet.cell(row=row, column=11, value=discipline_status['zade'])
    sheet.cell(row=row, column=12, value=foto_path)

    # Вставляем изображение в ячейку
    img = openpyxl.drawing.image.Image(foto_path)
    img.anchor = f"L{row}"
    img.width = 310  # Установка ширины изображения
    img.height = 230  # Установка высоты изображения
    sheet.add_image(img)
    img_width_pixels = 310
    img_height_pixels = 230
    img_width_points = img_width_pixels * 0.75
    img_height_points = img_height_pixels * 0.75
    sheet.row_dimensions[row].height = img_height_points
    sheet.column_dimensions['L'].width = img_width_points

    wb.save(excel_file)
    wb.close()
